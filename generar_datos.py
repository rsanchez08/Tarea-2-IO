"""
generar_datos.py
------------------------------------------------------------------
Construye 'datos_campo.xlsx' a partir del archivo de campo del grupo:
    datos_85_pasajeros.xlsx   (hoja 'observaciones')

Sistema : abordaje de pasajeros en la puerta del bus, parada de La Canada
          hacia Alajuela.  Servidor = la puerta.
Fecha   : se fija 2026-08-31 para todas las filas (pedido del grupo).
Servidor: se fija "Servidor 1" para todas las filas (pedido del grupo).

En la hoja 'observaciones', las 4 variables derivadas se escriben como
FORMULAS DE EXCEL (inciso 8 del enunciado), calculadas a partir de las
columnas de hora:
    interarribo_min     = (hora_llegada[fila] - hora_llegada[fila-1]) * 1440
    tiempo_espera_min   = (hora_inicio_servicio - hora_llegada) * 1440
    tiempo_servicio_min = (hora_fin_servicio - hora_inicio_servicio) * 1440
    tiempo_sistema_min  = (hora_fin_servicio - hora_llegada) * 1440
(1440 = minutos por dia; Excel guarda las fechas-hora como numero de dias)
------------------------------------------------------------------
"""

import pandas as pd
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter

ORIGEN   = "datos_85_pasajeros.xlsx"   # datos de campo ORIGINALES del grupo (ruta relativa)
FECHA    = "2026-08-31"
SERVIDOR = "Servidor 1"
INTERVALO_CONTEO_MIN = 10
VENTANA_INI, VENTANA_FIN = "06:00", "08:30"

# ----------------------------------------------------------------------
# 1. Leer el archivo del grupo (encabezado en la 4a fila -> header=3)
# ----------------------------------------------------------------------
src = pd.read_excel(ORIGEN, sheet_name="observaciones", header=3)
src = src.dropna(how="all").reset_index(drop=True)
print(f"Filas leidas del origen: {len(src)}")

# ----------------------------------------------------------------------
# 2. Fecha, servidor y fecha-hora COMPLETAS (como datetime, no texto)
# ----------------------------------------------------------------------
def a_dt(v):
    if pd.isna(v):
        return None
    if hasattr(v, "hour") and not hasattr(v, "year"):        # datetime.time
        hhmmss = v.strftime("%H:%M:%S")
    else:
        hhmmss = str(v)[-8:]
    return datetime.strptime(f"{FECHA} {hhmmss}", "%Y-%m-%d %H:%M:%S")

cols_hora = ["hora_llegada", "hora_inicio_servicio", "hora_fin_servicio",
            "hora_llegada_bus", "hora_salida_bus"]

df = pd.DataFrame()
df["id_observacion"]        = range(1, len(src) + 1)
df["fecha"]                 = FECHA
df["id_bus"]               = src["id_bus"].astype("Int64")
for c in cols_hora:
    df[c] = src[c].map(a_dt)
df["intervalo_observacion"] = src["intervalo_observacion"].astype(str)
df["servidor"]             = SERVIDOR
df["observaciones"]        = src["observaciones"].fillna("")

# ordenar por hora de llegada y renumerar
df = df.sort_values("hora_llegada").reset_index(drop=True)
df["id_observacion"] = range(1, len(df) + 1)

# numero de pasajeros que llegan en el mismo bloque de 10 min
df["numero_llegadas_intervalo"] = df.groupby("intervalo_observacion")["id_observacion"].transform("count")

# ----------------------------------------------------------------------
# 3. Variables derivadas: valores numericos (para CSV / resumen)
#    En el Excel iran como FORMULAS (paso 6).
# ----------------------------------------------------------------------
df["interarribo_min"]     = df["hora_llegada"].diff().dt.total_seconds().div(60).round(4)
df["tiempo_espera_min"]   = (df["hora_inicio_servicio"] - df["hora_llegada"]).dt.total_seconds().div(60).round(4)
df["tiempo_servicio_min"] = (df["hora_fin_servicio"] - df["hora_inicio_servicio"]).dt.total_seconds().div(60).round(4)
df["tiempo_sistema_min"]  = (df["hora_fin_servicio"] - df["hora_llegada"]).dt.total_seconds().div(60).round(4)

# orden final de columnas de la hoja 'observaciones'
COLS = ["id_observacion", "fecha", "id_bus",
        "hora_llegada", "hora_inicio_servicio", "hora_fin_servicio",
        "hora_llegada_bus", "hora_salida_bus",
        "intervalo_observacion", "numero_llegadas_intervalo", "servidor",
        "interarribo_min", "tiempo_espera_min", "tiempo_servicio_min", "tiempo_sistema_min",
        "observaciones"]
df = df[COLS]

# ----------------------------------------------------------------------
# 4. Hoja normalizada 'llegadas_intervalo' (15 bloques de 10 min)
# ----------------------------------------------------------------------
d0 = datetime.strptime(f"{FECHA} {VENTANA_INI}", "%Y-%m-%d %H:%M")
dfin = datetime.strptime(f"{FECHA} {VENTANA_FIN}", "%Y-%m-%d %H:%M")
n_int = int((dfin - d0).total_seconds() // 60 // INTERVALO_CONTEO_MIN)
llegadas = df["hora_llegada"]
filas_int = []
for k in range(n_int):
    b_ini = d0 + timedelta(minutes=k * INTERVALO_CONTEO_MIN)
    b_fin = b_ini + timedelta(minutes=INTERVALO_CONTEO_MIN)
    c = int(((llegadas >= b_ini) & (llegadas < b_fin)).sum())
    filas_int.append(dict(
        fecha=FECHA,
        intervalo_inicio=b_ini.strftime("%Y-%m-%d %H:%M:%S"),
        intervalo_observacion=f"{b_ini:%H:%M}-{b_fin:%H:%M}",
        duracion_min=INTERVALO_CONTEO_MIN,
        numero_llegadas_intervalo=c,
        servidor=SERVIDOR,
    ))
df_int = pd.DataFrame(filas_int)

# ----------------------------------------------------------------------
# 5. Hoja 'resumen' (agregado por columna) + hoja 'formulas' (documentacion)
# ----------------------------------------------------------------------
num_cols = ["interarribo_min", "tiempo_espera_min", "tiempo_servicio_min", "tiempo_sistema_min"]
res = pd.DataFrame({
    "columna":  num_cols + ["numero_llegadas_intervalo (15 bloques)"],
    "n":        [int(df[c].count()) for c in num_cols] + [len(df_int)],
    "suma":     [round(df[c].sum(), 3) for c in num_cols] + [int(df_int.numero_llegadas_intervalo.sum())],
    "promedio": [round(df[c].mean(), 4) for c in num_cols] + [round(df_int.numero_llegadas_intervalo.mean(), 4)],
    "minimo":   [round(df[c].min(), 3) for c in num_cols] + [int(df_int.numero_llegadas_intervalo.min())],
    "maximo":   [round(df[c].max(), 3) for c in num_cols] + [int(df_int.numero_llegadas_intervalo.max())],
    "desv_est": [round(df[c].std(ddof=1), 4) for c in num_cols] + [round(df_int.numero_llegadas_intervalo.std(ddof=1), 4)],
})
formulas_doc = pd.DataFrame({
    "variable": ["interarribo_min", "tiempo_espera_min", "tiempo_servicio_min", "tiempo_sistema_min"],
    "definicion": [
        "tiempo entre llegadas consecutivas",
        "desde la llegada hasta el inicio del servicio",
        "desde el inicio hasta la finalizacion del servicio",
        "desde la llegada hasta la finalizacion del servicio",
    ],
    "formula_excel (en la celda, fila r; datos desde la fila 2)": [
        "( D{r} - D{r-1} ) * 1440",
        "( E{r} - D{r} ) * 1440",
        "( F{r} - E{r} ) * 1440",
        "( F{r} - D{r} ) * 1440",
    ],
    "columnas": [
        "D = hora_llegada",
        "E = hora_inicio_servicio ; D = hora_llegada",
        "F = hora_fin_servicio ; E = hora_inicio_servicio",
        "F = hora_fin_servicio ; D = hora_llegada",
    ],
})

# ----------------------------------------------------------------------
# 6. Guardar Excel y, en la hoja 'observaciones', poner FORMULAS en L,M,N,O
# ----------------------------------------------------------------------
with pd.ExcelWriter("datos_campo.xlsx", engine="openpyxl",
                    datetime_format="YYYY-MM-DD HH:MM:SS") as xls:
    df.to_excel(xls, sheet_name="observaciones", index=False)
    df_int.to_excel(xls, sheet_name="llegadas_intervalo", index=False)
    res.to_excel(xls, sheet_name="resumen", index=False)
    formulas_doc.to_excel(xls, sheet_name="formulas", index=False)

    ws = xls.sheets["observaciones"]
    # letras de columna segun COLS (1-based)
    L = {name: get_column_letter(i + 1) for i, name in enumerate(COLS)}
    hl, hi, hf = L["hora_llegada"], L["hora_inicio_servicio"], L["hora_fin_servicio"]
    cL = {
        "interarribo_min":     lambda r: f"=({hl}{r}-{hl}{r-1})*1440",
        "tiempo_espera_min":   lambda r: f"=({hi}{r}-{hl}{r})*1440",
        "tiempo_servicio_min": lambda r: f"=({hf}{r}-{hi}{r})*1440",
        "tiempo_sistema_min":  lambda r: f"=({hf}{r}-{hl}{r})*1440",
    }
    n = len(df)
    for var, make in cL.items():
        col = L[var]
        for i in range(n):
            r = i + 2                       # fila 1 = encabezado; datos desde la fila 2
            if var == "interarribo_min" and i == 0:
                ws[f"{col}{r}"] = None      # la 1a fila no tiene "fila anterior"
            else:
                ws[f"{col}{r}"] = make(r)
            ws[f"{col}{r}"].number_format = "0.0000"

df.to_csv("datos_campo_observaciones.csv", index=False)
df_int.to_csv("datos_campo_llegadas_intervalo.csv", index=False)

print(f"Pasajeros            : {len(df)}")
print(f"fecha (todas)        : {df.fecha.unique().tolist()}")
print(f"servidor (todas)     : {df.servidor.unique().tolist()}")
print("Hoja 'observaciones': L=interarribo_min, M=tiempo_espera_min, "
      "N=tiempo_servicio_min, O=tiempo_sistema_min  ->  como FORMULAS")
print("Hoja 'formulas': documenta cada formula.")
print("\n--- resumen ---")
print(res.to_string(index=False))
